from xml.dom.minidom import parse
import xml.dom.minidom
import openpyxl



DOMTree=xml.dom.minidom.parse("go_obo.xml")
collection=DOMTree.documentElement
genes=collection.getElementsByTagName("term")
#collect all elements

wb=openpyxl.Workbook()
ws=wb.active

def get_childnode(gene_def):
    number0=gene_def.getElementsByTagName("is_a")
    number=number0.childNodes[0].data
    return(number)


def findgene(genes_list):
    ws.cell(1,1).value="GO_ID"
    ws.cell(1,2).value="Name"
    ws.cell(1,3).value="Definition"
    ws.cell(1,4).value="childnodes"
    count=2
    id_list=[]
    name_list=[]
    node_list=[]
    for gene in genes_list:
        current_gene_id0=gene.getElementsByTagName("id")[0]
        current_gene_id=current_gene_id0.childNodes[0].data
        current_gene_name0=gene.getElementsByTagName("name")[0]
        current_gene_name=current_gene_name0.childNodes[0].data
        gene_def=gene.getElementsByTagName("def")[0]
        gene_defstr=gene_def.getElementsByTagName("defstr")[0].childNodes[0].data
        if gene_defstr.find("autophagosome")!=0:
            ws.cell(row=count,column=1).value=current_gene_id
            ws.cell(row=count,column=2).value=current_gene_name
            ws.cell(row=count,column=3).value=current_gene_id
            ws.cell(row=count,column=4).value=current_gene_id
            count+=1
    wb.save("GO.xlsx")
        
